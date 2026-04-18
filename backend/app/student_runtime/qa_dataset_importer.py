from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.app.common.db import session_scope
from backend.chaoxing_db.models import QAFaqItem, QAFaqVariant

DEFAULT_DATASET_SHEET = "问答初始数据"


@dataclass(slots=True)
class DatasetRow:
    category: str | None
    canonical_question: str
    answer_text: str
    match_mode: str | None
    status: str
    similar_questions: list[str]
    source_file: str
    source_sheet: str


def load_excel_rows(file_path: str | Path, sheet_name: str = DEFAULT_DATASET_SHEET) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
        result.append(item)
    return result


def normalize_dataset_row(row: dict[str, Any], source_file: str, source_sheet: str) -> DatasetRow | None:
    canonical_question = _clean_text(
        row.get("标准问题(80字以内，不能为空)") or row.get("标准问题") or row.get("标准问题（80字以内，不能为空）")
    )
    answer_text = _clean_text(row.get("答案(2000字以内，不能为空)") or row.get("答案"))
    if not canonical_question or not answer_text:
        return None

    category = _clean_text(row.get("规则分类（30字以内，空为默认分类）") or row.get("规则分类"))
    match_mode = _clean_text(row.get("匹配模式"))
    status = _normalize_status(row.get("规则状态"))
    similar_questions = [
        item
        for item in (
            _clean_text(row.get("相似问法1(80字以内)") or row.get("相似问法1")),
            _clean_text(row.get("相似问法2(80字以内)") or row.get("相似问法2")),
            _clean_text(row.get("相似问法3(80字以内)") or row.get("相似问法3")),
        )
        if item
    ]
    return DatasetRow(
        category=category or None,
        canonical_question=canonical_question,
        answer_text=answer_text,
        match_mode=match_mode or None,
        status=status,
        similar_questions=similar_questions,
        source_file=source_file,
        source_sheet=source_sheet,
    )


def upsert_qa_dataset_rows(db: Session, rows: list[DatasetRow]) -> dict[str, int]:
    inserted = 0
    updated = 0
    for row in rows:
        item = db.query(QAFaqItem).filter(QAFaqItem.canonical_question == row.canonical_question).first()
        if item is None:
            item = QAFaqItem(
                category=row.category,
                canonical_question=row.canonical_question,
                answer_text=row.answer_text,
                match_mode=row.match_mode,
                status=row.status,
                source="organizer_dataset",
                source_file=row.source_file,
                source_sheet=row.source_sheet,
            )
            db.add(item)
            db.flush()
            inserted += 1
        else:
            item.category = row.category
            item.answer_text = row.answer_text
            item.match_mode = row.match_mode
            item.status = row.status
            item.source = "organizer_dataset"
            item.source_file = row.source_file
            item.source_sheet = row.source_sheet
            updated += 1
            db.flush()
            db.query(QAFaqVariant).filter(QAFaqVariant.faq_item_id == item.id).delete(synchronize_session=False)

        variants = [(row.canonical_question, "canonical", 0)]
        variants.extend((question, f"similar_{index}", index) for index, question in enumerate(row.similar_questions, start=1))
        for text, variant_type, sort_no in variants:
            db.add(
                QAFaqVariant(
                    faq_item_id=item.id,
                    variant_text=text,
                    variant_type=variant_type,
                    sort_no=sort_no,
                    is_active=row.status == "enabled",
                )
            )
    db.commit()
    return {"inserted": inserted, "updated": updated}


def import_dataset(db: Session, file_path: str | Path, sheet_name: str = DEFAULT_DATASET_SHEET) -> dict[str, int]:
    path = Path(file_path)
    normalized_rows: list[DatasetRow] = []
    for row in load_excel_rows(path, sheet_name=sheet_name):
        normalized = normalize_dataset_row(row, source_file=path.name, source_sheet=sheet_name)
        if normalized:
            normalized_rows.append(normalized)
    return upsert_qa_dataset_rows(db, normalized_rows)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u3000", " ").split()).strip()


def _normalize_status(value: Any) -> str:
    text = _clean_text(value).lower()
    if text in {"启用", "有效", "active", "enabled", "1", "true"}:
        return "enabled"
    if text in {"禁用", "无效", "inactive", "disabled", "0", "false"}:
        return "disabled"
    return "enabled"


if __name__ == "__main__":
    import sys

    dataset_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        r"D:\BaiduNetdiskDownload\10-a12基于泛雅平台的AI互动智课生成与实时问答\a12基于泛雅平台的AI互动智课生成与实时问答\教学问答数据集.xlsx"
    )
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_DATASET_SHEET
    with session_scope() as db:
        summary = import_dataset(db, dataset_path, sheet_name=sheet_name)
    print(summary)
