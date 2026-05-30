from __future__ import annotations

import json
import random
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from backend.app.common.vector_db import QAVectorChunk, vector_session_scope
from backend.app.student_runtime.qa_dashscope_client import DashScopeClient
from backend.app.student_runtime.qa_embedding_service import embed_texts


SAMPLE_SIZE = 60
SEED = 20260423
TOP_K = 5
REPORT_PATH = Path("output/eval/qa_e2e_generation_eval_60.json")
CHECKPOINT_PATH = Path("output/eval/qa_e2e_generation_eval_60.checkpoint.json")

REWRITE_TYPES = (
    "colloquial",
    "multi_turn",
    "inversion",
    "ellipsis",
    "pronoun_context",
    "noisy_distractor",
)


@dataclass(slots=True)
class FaqItem:
    item_id: int
    category: str
    canonical: str
    answer: str
    similars: list[str]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u3000", " ").split()).strip()


def find_dataset_path() -> Path:
    candidates = [path for path in Path("D:/BaiduNetdiskDownload").rglob("*.xlsx") if path.stat().st_size == 526501]
    if candidates:
        return candidates[0]
    all_xlsx = list(Path("D:/BaiduNetdiskDownload").rglob("*.xlsx"))
    if not all_xlsx:
        raise RuntimeError("No .xlsx dataset found under D:/BaiduNetdiskDownload")
    return all_xlsx[0]


def load_faq_items(path: Path) -> tuple[list[FaqItem], dict[int, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    next(rows, None)

    items: list[FaqItem] = []
    variant_to_item: dict[int, int] = {}
    variant_id = 1
    item_id = 1
    for row in rows:
        question = clean(row[1] if len(row) > 1 else None)
        answer = clean(row[2] if len(row) > 2 else None)
        if not question or not answer:
            continue
        similars: list[str] = []
        for column in (5, 6, 7):
            if len(row) > column:
                similar = clean(row[column])
                if similar:
                    similars.append(similar)
        for _ in [question, *similars]:
            variant_to_item[variant_id] = item_id
            variant_id += 1
        items.append(
            FaqItem(
                item_id=item_id,
                category=clean(row[0] if len(row) > 0 else None),
                canonical=question,
                answer=answer,
                similars=similars,
            )
        )
        item_id += 1
    workbook.close()
    return items, variant_to_item


def extract_term(question: str) -> str:
    text = clean(question)
    for prefix in ("什么是", "如何理解", "为什么", "怎么理解", "请解释", "解释"):
        if text.startswith(prefix):
            text = text[len(prefix) :]
            break
    for token in ("是什么", "的定义", "定义"):
        text = text.replace(token, "")
    text = re.sub(r"[？?。！!]", "", text).strip(" ：:，,。")
    return text[:28] if text else question[:28]


def build_query(item: FaqItem, rewrite_type: str) -> list[dict[str, str]]:
    term = extract_term(item.canonical)
    similar = item.similars[0] if item.similars else item.canonical
    if rewrite_type == "colloquial":
        return [{"role": "user", "content": f"老师，{term}这块我有点懵，能不能用简单点的话讲讲？"}]
    if rewrite_type == "multi_turn":
        return [
            {"role": "user", "content": "我刚才没完全听懂这个知识点。"},
            {"role": "assistant", "content": "你想问哪个概念？"},
            {"role": "user", "content": f"就是“{term}”，它具体是什么意思？"},
        ]
    if rewrite_type == "inversion":
        return [{"role": "user", "content": f"如果要解释清楚的话，{term}到底应该怎么定义？"}]
    if rewrite_type == "ellipsis":
        return [{"role": "user", "content": f"{term}呢？课堂里这个点怎么理解？"}]
    if rewrite_type == "pronoun_context":
        return [
            {"role": "user", "content": f"前面讲到“{term}”。"},
            {"role": "user", "content": "这个它主要指什么？有什么关键点？"},
        ]
    if rewrite_type == "noisy_distractor":
        return [{"role": "user", "content": f"不是问平台登录、作业提交，也不是问考试安排；我只想问知识点：{similar}"}]
    raise ValueError(f"Unknown rewrite type: {rewrite_type}")


def render_messages(messages: list[dict[str, str]]) -> str:
    return "\n".join(f"{message['role']}: {message['content']}" for message in messages)


def build_generation_prompt(query: str, contexts: list[dict[str, Any]]) -> str:
    context_lines = []
    for index, context in enumerate(contexts, start=1):
        context_lines.append(
            f"[{index}] 问题：{context['question']}\n答案：{context['answer']}"
        )
    return (
        "你是课程智能问答助手。请只基于给定的召回资料回答学生问题；"
        "如果资料中没有足够依据，请明确说明信息不足，不要编造。\n\n"
        f"学生问题/对话：\n{query}\n\n"
        "召回资料：\n"
        + "\n\n".join(context_lines)
        + "\n\n请给出简洁、准确、自然的中文回答。"
    )


def parse_judge_json(raw: str) -> dict[str, Any]:
    text = raw.strip()
    if text.startswith("```"):
        match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.S)
        if match:
            text = match.group(1)
    match = re.search(r"\{.*\}", text, re.S)
    if match:
        text = match.group(0)
    return json.loads(text)


def build_judge_prompt(query: str, expected_answer: str, generated_answer: str) -> str:
    return f"""
你是严格的教学问答评测员。请对模型回答评分，输出 JSON，不要输出额外文本。

评分规则：
- accuracy: 0-100，回答是否覆盖标准答案核心事实，是否有事实错误。
- relevance: 0-100，回答是否直接回应学生问题，是否跑题。
- fluency: 0-100，中文表达是否通顺自然、结构清楚。
- correct: true/false，当 accuracy >= 80 且 relevance >= 75 时为 true。

学生问题：
{query}

标准答案：
{expected_answer}

模型回答：
{generated_answer}

输出格式：
{{"accuracy": 0, "relevance": 0, "fluency": 0, "correct": false, "reason": "一句话说明主要扣分点"}}
""".strip()


def build_batch_judge_prompt(cases: list[dict[str, Any]]) -> str:
    payload = [
        {
            "caseId": case["caseId"],
            "query": case["query"],
            "expectedAnswer": case["expectedAnswer"],
            "generatedAnswer": case["generatedAnswer"],
        }
        for case in cases
    ]
    return (
        "你是严格的教学问答评测员。请对每个模型回答评分，只输出 JSON 数组，不要输出额外文本。\n\n"
        "评分规则：\n"
        "- accuracy: 0-100，回答是否覆盖标准答案核心事实，是否有事实错误。\n"
        "- relevance: 0-100，回答是否直接回应学生问题，是否跑题。\n"
        "- fluency: 0-100，中文表达是否通顺自然、结构清楚。\n"
        "- correct: true/false，当 accuracy >= 80 且 relevance >= 75 时为 true。\n\n"
        "输入样本 JSON：\n"
        f"{json.dumps(payload, ensure_ascii=False, indent=2)}\n\n"
        '输出格式：[{ "caseId": 1, "accuracy": 0, "relevance": 0, "fluency": 0, "correct": false, "reason": "一句话说明主要扣分点" }]'
    )


def parse_batch_judge_json(raw: str) -> list[dict[str, Any]]:
    text = raw.strip()
    if text.startswith("```"):
        match = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", text, re.S)
        if match:
            text = match.group(1)
    match = re.search(r"\[.*\]", text, re.S)
    if match:
        text = match.group(0)
    return json.loads(text)


def main() -> None:
    dataset_path = find_dataset_path()
    items, variant_to_item = load_faq_items(dataset_path)
    item_by_id = {item.item_id: item for item in items}

    random.seed(SEED)
    sampled_items = random.sample(items, SAMPLE_SIZE)
    test_cases = []
    for index, item in enumerate(sampled_items):
        rewrite_type = REWRITE_TYPES[index % len(REWRITE_TYPES)]
        messages = build_query(item, rewrite_type)
        test_cases.append(
            {
                "caseId": index + 1,
                "rewriteType": rewrite_type,
                "messages": messages,
                "query": render_messages(messages),
                "expectedItemId": item.item_id,
                "canonical": item.canonical,
                "expectedAnswer": item.answer,
                "category": item.category,
            }
        )

    embedding_started = time.perf_counter()
    embeddings = embed_texts([case["query"] for case in test_cases], text_type="query", batch_size=10)
    embedding_latency_ms = (time.perf_counter() - embedding_started) * 1000

    client = DashScopeClient()
    results = []
    retrieval_latencies = []
    generation_latencies = []
    judge_latencies = []
    retrieval_top1 = retrieval_top3 = retrieval_top5 = 0

    with vector_session_scope() as vector_db:
        for case, embedding in zip(test_cases, embeddings):
            retrieval_started = time.perf_counter()
            rows = (
                vector_db.query(QAVectorChunk.source_id, QAVectorChunk.chunk_text)
                .filter(QAVectorChunk.source_type == "faq_variant", QAVectorChunk.is_active == True)  # noqa: E712
                .order_by(QAVectorChunk.embedding.cosine_distance(embedding))
                .limit(TOP_K)
                .all()
            )
            retrieval_latency = (time.perf_counter() - retrieval_started) * 1000
            retrieval_latencies.append(retrieval_latency)

            retrieved_variant_ids = [int(row[0]) for row in rows]
            retrieved_item_ids = [variant_to_item.get(variant_id) for variant_id in retrieved_variant_ids]
            retrieved_items = []
            seen_item_ids = set()
            for item_id in retrieved_item_ids:
                if not item_id or item_id in seen_item_ids:
                    continue
                item = item_by_id.get(item_id)
                if not item:
                    continue
                seen_item_ids.add(item_id)
                retrieved_items.append(
                    {
                        "itemId": item.item_id,
                        "question": item.canonical,
                        "answer": item.answer,
                        "category": item.category,
                    }
                )

            expected_id = case["expectedItemId"]
            top1_hit = bool(retrieved_item_ids[:1] and retrieved_item_ids[0] == expected_id)
            top3_hit = expected_id in retrieved_item_ids[:3]
            top5_hit = expected_id in retrieved_item_ids[:5]
            retrieval_top1 += int(top1_hit)
            retrieval_top3 += int(top3_hit)
            retrieval_top5 += int(top5_hit)

            generation_prompt = build_generation_prompt(case["query"], retrieved_items)
            generation_started = time.perf_counter()
            generated = client.chat_completion(prompt=generation_prompt)["text"].strip()
            generation_latency = (time.perf_counter() - generation_started) * 1000
            generation_latencies.append(generation_latency)

            results.append(
                {
                    **case,
                    "retrievedVariantIds": retrieved_variant_ids,
                    "retrievedItemIds": retrieved_item_ids,
                    "retrievalTop1Hit": top1_hit,
                    "retrievalTop3Hit": top3_hit,
                    "retrievalTop5Hit": top5_hit,
                    "retrievedContexts": retrieved_items,
                    "generatedAnswer": generated,
                    "judge": None,
                    "retrievalLatencyMs": round(retrieval_latency, 2),
                    "generationLatencyMs": round(generation_latency, 2),
                    "judgeLatencyMs": None,
                }
            )
            CHECKPOINT_PATH.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
            print(
                json.dumps(
                    {
                        "progress": f"{len(results)}/{len(test_cases)}",
                        "caseId": case["caseId"],
                        "rewriteType": case["rewriteType"],
                        "retrievalTop1Hit": top1_hit,
                        "generationLatencyMs": round(generation_latency, 2),
                    },
                    ensure_ascii=False,
                ),
                flush=True,
            )

    for start in range(0, len(results), 8):
        batch = results[start : start + 8]
        judge_prompt = build_batch_judge_prompt(batch)
        judge_started = time.perf_counter()
        judge_raw = client.chat_completion(prompt=judge_prompt)["text"]
        judge_latency = (time.perf_counter() - judge_started) * 1000
        judge_latencies.append(judge_latency)
        try:
            judged = parse_batch_judge_json(judge_raw)
        except Exception:
            judged = [
                {
                    "caseId": item["caseId"],
                    "accuracy": 0,
                    "relevance": 0,
                    "fluency": 0,
                    "correct": False,
                    "reason": f"Batch judge JSON parse failed: {judge_raw[:200]}",
                }
                for item in batch
            ]
        judge_by_case_id = {int(item.get("caseId")): item for item in judged if item.get("caseId") is not None}
        for item in batch:
            item["judge"] = judge_by_case_id.get(
                int(item["caseId"]),
                {
                    "accuracy": 0,
                    "relevance": 0,
                    "fluency": 0,
                    "correct": False,
                    "reason": "Missing judge result in batch output.",
                },
            )
            item["judgeLatencyMs"] = round(judge_latency, 2)
        CHECKPOINT_PATH.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps({"judgeProgress": f"{min(start + 8, len(results))}/{len(results)}"}, ensure_ascii=False), flush=True)

    def average(values: list[float]) -> float:
        return round(sum(values) / len(values), 2) if values else 0.0

    by_type = {}
    for rewrite_type in REWRITE_TYPES:
        subset = [result for result in results if result["rewriteType"] == rewrite_type]
        by_type[rewrite_type] = {
            "sampleSize": len(subset),
            "retrievalTop1HitRate": round(sum(item["retrievalTop1Hit"] for item in subset) / len(subset), 4),
            "retrievalTop3HitRate": round(sum(item["retrievalTop3Hit"] for item in subset) / len(subset), 4),
            "answerAccuracyRate": round(sum(bool(item["judge"].get("correct")) for item in subset) / len(subset), 4),
            "avgAccuracyScore": average([float(item["judge"].get("accuracy", 0)) for item in subset]),
            "avgRelevanceScore": average([float(item["judge"].get("relevance", 0)) for item in subset]),
            "avgFluencyScore": average([float(item["judge"].get("fluency", 0)) for item in subset]),
        }

    report = {
        "metric": "end_to_end_retrieval_generation_quality",
        "definition": "End-to-end answer is correct when an LLM judge marks the generated answer as correct against the gold answer (accuracy >= 80 and relevance >= 75).",
        "datasetPath": str(dataset_path),
        "sampleSize": len(results),
        "randomSeed": SEED,
        "rewriteTypes": list(REWRITE_TYPES),
        "retrievalTop1HitRate": round(retrieval_top1 / len(results), 4),
        "retrievalTop3HitRate": round(retrieval_top3 / len(results), 4),
        "retrievalTop5HitRate": round(retrieval_top5 / len(results), 4),
        "answerAccuracyRate": round(sum(bool(item["judge"].get("correct")) for item in results) / len(results), 4),
        "answerAccuracyPercent": round(sum(bool(item["judge"].get("correct")) for item in results) / len(results) * 100, 2),
        "avgAccuracyScore": average([float(item["judge"].get("accuracy", 0)) for item in results]),
        "avgRelevanceScore": average([float(item["judge"].get("relevance", 0)) for item in results]),
        "avgFluencyScore": average([float(item["judge"].get("fluency", 0)) for item in results]),
        "avgRetrievalLatencyMs": average(retrieval_latencies),
        "avgGenerationLatencyMs": average(generation_latencies),
        "avgJudgeLatencyMs": average(judge_latencies),
        "embeddingLatencyMs": round(embedding_latency_ms, 2),
        "byRewriteType": by_type,
        "failureCases": [item for item in results if not bool(item["judge"].get("correct"))],
        "caseResults": results,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
    }
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    summary = {key: value for key, value in report.items() if key not in {"caseResults", "failureCases"}}
    summary["failureCount"] = len(report["failureCases"])
    summary["reportPath"] = str(REPORT_PATH.resolve())
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
